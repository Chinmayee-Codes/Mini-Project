from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Define weeks and corresponding days
weeks = {
    "Week 1": ["27 May (Mon)", "28 May (Tue)", "29 May (Wed)", "30 May (Thu)", "31 May (Fri)", "1 June (Sat)", "2 June (Sun)"],
    "Week 2": ["3 June (Mon)", "4 June (Tue)", "5 June (Wed)", "6 June (Thu)", "7 June (Fri)", "8 June (Sat)", "9 June (Sun)"],
    "Week 3": ["10 June (Mon)", "11 June (Tue)", "12 June (Wed)", "13 June (Thu)", "14 June (Fri)", "15 June (Sat)", "16 June (Sun)"]
}

# Time slots (same for all weeks)
time_slots = [
    "4:00 – 4:30 AM", "4:30 – 6:00 AM", "6:00 – 8:00 AM", "9:00 – 10:30 AM", "10:30 – 12:00 PM",
    "12:00 – 1:00 PM", "1:00 – 2:30 PM", "2:30 – 3:00 PM", "3:00 – 4:00 PM", "4:00 – 5:30 PM",
    "5:30 – 6:30 PM", "6:30 – 7:45 PM", "7:45 – 8:15 PM", "8:15 – 9:00 PM"
]

# Paste your week_data here based on the structured format above
week_data = {
    "Week 1": [
        ["Morning Routine"]*7,
        ["SQL – SELECT, WHERE, GROUP BY", "SQL Joins – INNER, LEFT, RIGHT", "SQL Window Functions", "SQL Optimization", "Data Modeling Concepts", "Weekly Revision – SQL + ETL", "Light Review of Week"],
        ["Gym + Freshen Up"]*7,
        ["Intro to ETL + Apache NiFi", "NiFi Processors – Data Ingestion", "Apache Airflow Intro", "Airflow DAGs", "Python (Pandas for DE)", "ML – Ensemble Methods", "Project Work or Hackathon"],
        ["GenAI Basics + Prompting", "Prompt Engineering Basics", "GenAI – Foundation Models", "GenAI – Fine-Tuning Basics", "Prompt Tips (LLMs)", "Resume + LinkedIn Review", "Review + Planning for Week"],
        ["Lunch + Nap"]*7,
        ["Consulting – What does a consultant do?", "ML – Decision Trees", "ML – KNN, Naïve Bayes", "ML – SVM", "ML – Evaluation Metrics", "Mock Interview: DE/ML", "Rest or Networking Outreach"],
        ["Tea Break"]*7,
        ["Tuition"]*6 + ["—"],
        ["ML – Linear & Logistic Regression", "Consulting – Market Entry", "Consulting – Profitability", "Consulting – Case Practice #1", "Consulting – Behavioral Q&A", "—", "—"],
        ["Guesstimates – Market Size, Estimations", "—", "—", "—", "—", "—", "—"],
        ["—"]*7,
        ["Dinner"]*7,
        ["Recap + Flashcards"]*6 + ["Light Reflection"]
    ],
    "Week 2": [
        ["Morning Routine"]*7,
        ["Advanced SQL: CTEs, CASE WHEN", "Kafka Basics", "LangChain Basics", "Python Data Pipelines", "HuggingFace Intro", "Weekly Revision: Cloud + ML", "Hackathon or Kaggle Participation"],
        ["Gym + Freshen Up"]*7,
        ["Kafka Stream Processing", "Snowflake Overview", "Redshift Basics", "Cloud: GCP & BigQuery", "AWS S3 + Athena Basics", "Resume Refinement", "Self Review + Networking Outreach"],
        ["LLM APIs (OpenAI)", "LangChain Agents", "Prompt Engineering for Custom Apps", "Python for DE – File I/O, APIs", "Apache Beam Overview", "Mock Interview – Consulting Case", "—"],
        ["Lunch + Nap"]*7,
        ["ML – Clustering (K-Means)", "ML – Dimensionality Reduction (PCA)", "ML – Feature Engineering", "ML – Cross Validation", "ML – Hyperparameter Tuning", "—", "—"],
        ["Tea Break"]*7,
        ["Tuition", "Tuition (shifted 6:30 PM from June 4)", "Tuition", "Tuition", "Tuition", "Tuition", "—"],
        ["Consulting – Case Practice #2", "Consulting – Excel Practice", "Consulting – Market Sizing", "Consulting – PowerPoint Slide Making", "Consulting – Behavioral Round Prep", "—", "—"],
        ["—"]*7,
        ["—", "Tuition (new timing starts)", "—", "—", "—", "—", "—"],
        ["Dinner"]*7,
        ["Recap + Flashcards"]*6 + ["Rest and Planning"]
    ],
    "Week 3": [
        ["Morning Routine"]*7,
        ["Data Lake vs Warehouse", "Batch vs Stream Processing", "BigQuery Hands-On Lab", "SQL + GCP Revision", "DE Project Review", "Mock Interviews (Technical + Behavioral)", "Weekly Wrap-up"],
        ["Gym + Freshen Up"]*7,
        ["ETL Optimization", "Data Pipeline Project Planning", "Python + Airflow Project", "Full DE Mini Project Execution", "GitHub Upload + Documentation", "Case Study Jam Session", "Rest + Planning Next Phase"],
        ["Fine-Tuning OpenAI GPT", "HuggingFace Transformers Hands-on", "Prompt Templates + Evaluation", "Final Prompt Engineering Recap", "ML Project Review", "Community Post or Blog Draft", "—"],
        ["Lunch + Nap"]*7,
        ["ML – Model Deployment (Flask)", "ML – End-to-End Mini Project Setup", "ML – Model Monitoring", "ML – Project Report Writing", "Final Resume Review + Application Sprint", "—", "—"],
        ["Tea Break"]*7,
        ["Tuition"]*6 + ["—"],
        ["Consulting – Advanced Case Practice", "Consulting – Domain Knowledge (Tech)", "Consulting – Final Case Practice", "Consulting – Practice with Peer", "—", "—", "—"],
        ["—"]*7,
        ["Tuition (June 4 onwards)", "Tuition", "Tuition", "Tuition", "Tuition", "Tuition", "—"],
        ["Dinner"]*7,
        ["Recap + Flashcards"]*6 + ["Rest and Reflection"]
    ]
}


# Due to space, if you need the full dataset, let me know and I’ll give you `week_data` separately.

# Create workbook and fill in sheets
wb = Workbook()
for week, days in weeks.items():
    ws = wb.create_sheet(title=week)
    ws.append(["Time"] + days)
    for idx, slot in enumerate(time_slots):
        row = [slot] + week_data[week][idx]
        ws.append(row)
    for col in range(1, len(days) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 25

# Remove default sheet
default_sheet = wb["Sheet"]
wb.remove(default_sheet)

# Save workbook
wb.save("3_Week_Study_Schedule.xlsx")
